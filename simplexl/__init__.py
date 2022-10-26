__version__ = '0.1.0'

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class CreateExcel:

    def __init__(self, excel_name, header_columns, row_data, sheet_name="sheet1", sheet_index=0):

        self.header_columns = header_columns
        self.sheet_name = sheet_name
        self.sheet_index = sheet_index
        self.excel_name = excel_name
        self.row_data = row_data
        self.create_sheet()

    def create_sheet(self):
        wb = Workbook()
        wb.remove(wb.active)
        work_sheet = wb.create_sheet(
            title=self.sheet_name,
            index=self.sheet_index
        )
        self.populate_headers_to_excel(work_sheet)
        self.populate_row_data_to_excel(work_sheet)
        wb.save(self.excel_name)

    def populate_headers_to_excel(self, work_sheet):
        try:
            row_num = 1
            for col_num, col_data in enumerate(self.header_columns, 1):
                cell = work_sheet.cell(row=row_num, column=col_num)
                cell.value = str(col_data.get('name')).upper()
                cell.font = Font(
                    name=col_data.get('font_name', 'Calibri'),
                    bold=True,
                    color=col_data.get('font_color', 'FFFFFF'),
                )
                cell.alignment = Alignment(
                    vertical=col_data.get('alignment', 'center'),
                    horizontal=col_data.get('alignment', 'center'),
                    wrap_text=False
                )

                cell.fill = PatternFill(
                    start_color=col_data.get('bg_color', '5FABE6'),
                    end_color=col_data.get('bg_color', '5FABE6'),
                    fill_type='solid'
                )
                column_letter = get_column_letter(col_num)
                column_dimensions = work_sheet.column_dimensions[column_letter]
                column_dimensions.width = col_data.get('width', 20)
        except Exception as err:
            return err

    def populate_row_data_to_excel(self, work_sheet):
        try:
            row_num = 1
            for row in self.row_data:
                row_num += 1
                for col_num in range(1, len(row)+1):
                    cell = work_sheet.cell(row=row_num, column=col_num)
                    cell.value = row[col_num-1]
                    cell.alignment = Alignment(vertical='top', wrap_text=False)
        except Exception as err:
            return err
